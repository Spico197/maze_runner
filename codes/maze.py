# !python3
# coding=utf-8

"""
This is a python3 and pygame based maze runner game lib
for the Programming Course Design.

@author: Spico
@date: 2017/07/06
@repo: github.com/Spico197/maze_runner
"""

# ------------------ Lib Import ------------------ #
import pygame
import math
import random
from openpyxl import load_workbook

# ------------------ Color Definition ------------------ #
BLACK = (0, 0, 0)
WHITE = (255, 255, 255)
GREEN = (0, 255, 0)
RED = (255, 0, 0)
WARM_GREY = (239, 240, 220)
WARM_ORANGE = (242, 159, 63)
WARM_YELLOW = (255, 233, 87)

# ------------------ Global Variables Definition ------------------ #
# Screen dimensions
ROAD_THICKNESS = 50
WALL_THICKNESS = 6

SCREEN_WIDTH = 800
SCREEN_HEIGHT = 600

blood = 100

done = False  # main-loop flag

# ---------------- Resource Data Path ------------------ #
FONT_PATH = "../data/font/msyh.ttc"
TEACHER_IMAGE_PATH = "../data/pic/ds_30.jpg"
PLAYER_IMAGE_PATH = "../data/pic/mumu_30.jpg"
BACKGROUND_MUSIC_PATH = "../data/music/bgm_zalababa.mp3"
PRINCESS_IMAGE_PATH = "../data/pic/xiaoxiao_50.jpg"
QUESTION_WORKBOOK_PATH = "../data/question/question_cn.xlsx"
RIGHT_ANSWER_MUSIC_PATH = "../data/music/right.mp3"
WRONG_ANSWER_MUSIC_PATH = "../data/music/wrong.mp3"
LOSE_GAME_MUSIC_PATH = "../data/music/lose.mp3"
VICTORY_MUSIC_PATH = "../data/music/victory.ogg"


# ------------------ Class Definition ------------------ #

class Maze(object):
    """
    maze class
    """
    def __init__(self):
        """
        constructor init
        """
        self.neighbor = []
        self.visited = []
        self.unvisited = []
        self.wall_table = []
        self.road_mat = []

    def maze_generate(self, road_list, wall_list):
        """
        generate the road, wall from road list
        @param road_list: road sprite group
        @param wall_list: wall sprite group
        @return: road_list, wall_list
        """
        rows = len(self.road_mat)
        cols = len(self.road_mat[0])

        for row in range(rows):    # row&col start from 0 to cols-1
            for col in range(cols):
                if self.road_mat[row][col] == 1:
                    road = Road(ROAD_THICKNESS * col, ROAD_THICKNESS * row)
                    road_list.add(road)
                # get in all the matrix points
                self.unvisited.append([row, col])

        # wall maze drawing
        for row_inv in range(1, rows):
            for col_inv in range(0, cols):
                wall = Wall(col_inv * ROAD_THICKNESS,
                            row_inv * ROAD_THICKNESS - WALL_THICKNESS / 2, 1)
                wall_list.add(wall)
                self.wall_table.append(wall)
        for col_inv in range(1, cols):
            for row_inv in range(0, rows):
                wall = Wall(col_inv * ROAD_THICKNESS - WALL_THICKNESS / 2,
                            row_inv * ROAD_THICKNESS, 2)
                wall_list.add(wall)
                self.wall_table.append(wall)

        return road_list, wall_list

    def isexit(self, pos_now):     # row&col start from 0
        """
        test if the pos_now is the exit
        @param pos_now: the position remaining to be tested
        @return: flag-the bool variable
        """

        flag = False
        row = pos_now[0]
        col = pos_now[1]
        # find the exit
        exit_row = 0
        exit_col = 0

        for row_mat in range(len(self.road_mat)):
            for col_mat in range(len(self.road_mat[0])):
                if self.road_mat[row_mat][col_mat] == -1:
                    exit_row = row_mat
                    exit_col = col_mat

        if exit_row == row and exit_col == col:
            flag = True

        return flag

    def dfs_maze_generate(self, start_pos, wall_list):
        """
        dfs algorithm to generate the wall group
        @param start_pos: the position that mumu is ready to go
        @param wall_list: wall group remaining to change
        @return: if the maze is good to run, return wall_list,
                 else return False
        """
        self.unvisited.remove(start_pos)
        # maze.visited add the start_pos and remove the start_pos in unvisited
        self.visited.append(start_pos)
        current_stack = [start_pos]

        find = False
        while self.unvisited:
            next_pos = self.neighbor_select(start_pos)

            if not next_pos:
                if current_stack:
                    start_pos = current_stack.pop()
                else:
                    start_pos = self.unvisited[
                        random.randrange(0, len(self.unvisited))]
                continue

            current_stack.append(next_pos)

            self.road_mat[next_pos[0]][next_pos[1]] = 0
            self.visited.append(next_pos)
            self.unvisited.remove(next_pos)
            wall_list = self.wall_break(start_pos, next_pos, wall_list)

            if not wall_list:
                return False

            if self.isexit(next_pos):
                find = True

            start_pos = next_pos
            # dfs_maze_generate(next_pos, maze, maze_matrix)
        if find:
            return wall_list
        else:
            return False

    def neighbor_select(self, pos_now):
        """
        select a neighbor near pos_now
        @param pos_now: the current position
        @return: the neighbor position
        """

        row = len(self.road_mat)
        col = len(self.road_mat[0])
        row_now = pos_now[0]
        col_now = pos_now[1]
        # judge the up&down tag comparision
        tag_up_down = row_now * col + col_now
        # judge the left&right tag comparision
        tag_left_right = col_now * row + row_now
        neighbor = []

        if tag_up_down + col <= row * col + col \
                and [row_now + 1, col_now] in self.unvisited:
            neighbor.append([row_now + 1, col_now])
        if tag_up_down - col >= 0 and [row_now - 1, col_now] in self.unvisited:
            neighbor.append([row_now - 1, col_now])
        if tag_left_right + row <= row * col + col \
                and [row_now, col_now + 1] in self.unvisited:
            neighbor.append([row_now, col_now + 1])
        if tag_left_right - row >= 0 \
                and [row_now, col_now - 1] in self.unvisited:
            neighbor.append([row_now, col_now - 1])

        if not neighbor:
            return False

        return neighbor[random.randrange(0, len(neighbor))]

    def wall_break(self, pos1, pos2, wall_list):
        """
        break the wall between to points
        @param pos1: first position
        @param pos2: second position
        @param wall_list: the wall group
        @return: if handy, return wall_list, else return False
        """
        rows = len(self.road_mat)
        cols = len(self.road_mat[0])
        location = 0

        # left & right border
        if pos1[0] == pos2[0] and math.fabs(pos1[1] - pos2[1]) == 1:
            if pos1[1] - pos2[1] == -1:  # pos1 at the left
                location = (rows - 1) * cols + pos1[1] * rows + pos1[0]
            elif pos1[1] - pos2[1] == 1:
                location = (rows - 1) * cols + pos2[1] * rows + pos2[0]
        elif pos1[1] == pos2[1] and math.fabs(pos1[0] - pos2[0]) == 1:
            if pos1[0] - pos2[0] == 1:  # pos1 at the bottom
                location = pos2[0] * cols + pos2[1]
            elif pos1[0] - pos2[0] == -1:
                location = pos1[0] * cols + pos1[1]
        try:
            wall_list.remove(self.wall_table[location])
            return wall_list
        except Exception as e:
            print("wall_remove error!" + str(Exception) + str(e))
            return False


class Road(pygame.sprite.Sprite):
    """
    @brief: road class inherited from sprite and image loaded
    @param: x: x-direction position
    @param: y: y-direction position
    """

    def __init__(self, x, y):  # constructor function
        """
        constructor init function
        @param x: the road x position
        @param y: the road y position
        """
        super().__init__()  # parents' constructor call

        self.image = pygame.Surface((ROAD_THICKNESS, ROAD_THICKNESS))
        self.image.fill(WARM_ORANGE)
        self.rect = self.image.get_rect()
        self.rect.x = x
        self.rect.y = y


class Wall(pygame.sprite.Sprite):
    """
    @brief: wall class inherited from sprite and image loaded
    @param: x: x-direction position
    @param: y: y-direction position
    """

    def __init__(self, x, y, mode):  # constructor function
        """
        constructor function
        @param x: x position
        @param y: y position
        @param mode: mode 1: the standing one; mode 2: the lying one
        """
        super().__init__()  # parents' constructor call

        if mode == 1:   # mode = 1, lying down; mode = 2, stand up.
            self.image = pygame.Surface((ROAD_THICKNESS, WALL_THICKNESS))
        if mode == 2:
            self.image = pygame.Surface((WALL_THICKNESS, ROAD_THICKNESS))
        self.image.fill(BLACK)
        self.rect = self.image.get_rect()
        self.rect.x = x
        self.rect.y = y


class Player(pygame.sprite.Sprite):
    """
    @brief: the Player class inherited from sprite and image loaded
    """

    change_x = 0    # speed
    change_y = 0

    def __init__(self, x, y):
        """
        constructor function
        @param x: x-position
        @param y: y-position
        """

        super().__init__()  # inherited

        self.image = pygame.image.load(PLAYER_IMAGE_PATH)
        # image/rect are the properties of the sprite class
        self.rect = self.image.get_rect()   # top left corner location
        self.rect.y = y
        self.rect.x = x

    def changespeed(self, x, y):
        """
        change the speed of the player
        @param x: x-speed
        @param y: y-speed
        """

        self.change_x += x
        self.change_y += y

    def reset_speed(self):
        self.change_x = 0
        self.change_y = 0

    def move(self, walls):
        """
        move the player
        @param walls: the wall group that player cannot craw out
        @return: none
        """

        # first left/right then up/down

        # moving left/right
        self.rect.x += self.change_x

        if self.rect.x < 0:
            self.rect.left = 0
        if self.rect.right > SCREEN_WIDTH:
            self.rect.right = SCREEN_WIDTH

        # create the interact list
        block_hit_list = pygame.sprite.spritecollide(self, walls, False)

        # set the limitation of the player
        for block in block_hit_list:
            # moving right, the hit obj's left side is the player's right side
            if self.change_x > 0:
                self.rect.right = block.rect.left
            else:
                # moving left, the road's right is the player's left
                self.rect.left = block.rect.right

        # moving up/down
        self.rect.y += self.change_y

        if self.rect.y < 0:
            self.rect.top = 0
        if self.rect.bottom > SCREEN_HEIGHT:
            self.rect.bottom = SCREEN_HEIGHT

        # create hit list
        block_hit_list = pygame.sprite.spritecollide(self, walls, False)

        # limitation in the up&down direction
        for block in block_hit_list:
            if self.change_y > 0:
                # NOTICE: it's moving down
                self.rect.bottom = block.rect.top
            else:
                self.rect.top = block.rect.bottom


class Teacher(pygame.sprite.Sprite):
    def __init__(self):
        """
        constructor function
        """
        super().__init__()
        self.image = pygame.image.load(TEACHER_IMAGE_PATH).convert()
        self.rect = self.image.get_rect()  # top-left corner location


# ------------------ Function Definition ------------------ #
def display_texts_page(screen, texts, flip=True):
    screen.fill(WARM_GREY)
    font_wel = pygame.font.Font(FONT_PATH, 20)
    for text_idx, text in enumerate(texts):
        text_wel = font_wel.render(text, True, BLACK)
        screen.blit(text_wel, [50, 50 + text_idx * 30])
    if flip:
        pygame.display.flip()


def welcome(screen_wel):
    """
    print the welcome information
    @param screen_wel: the screen which is ready to make an output
    @return: none
    """
    texts = [
        "故事发生在名为大数据的王国，",
        "这里的人们以知识和技术为尊，人们享受着新兴技术",
        "给他们带来的生活便利",
        "有一天，王宫中传来了国王的叹息声，",
        "原来国王唯一的女儿小小在闯期末考核时发生了意外，",
        "被困在知识迷岛里面，国王不但担忧着女儿的安危，",
        "更担忧女儿无法继承他的王位，思虑多时，国王最终",
        "决定，贴榜昭告天下，凡是能救出公主的，就可以成为",
        "王国的驸马。万千勇士收到消息便不及待地踏上了征程，",
        "而这当中就有一位名唤木木的少年……",
    ]
    display_texts_page(screen_wel, texts)
    pygame.time.wait(3000)

    texts = [
        "勇士木木来到了知识迷岛，却发现这座岛比他想象的要",
        "复杂得多，整座岛是一个非常大的迷宫，道路错综复杂，",
        "更麻烦的是重要的关卡还有擅长不同学科的老师的把守，",
        "要想让老师放路，就只能乖乖的通过老师的考试，通过",
        "考试就可以得到迷宫道路中的提示，找到公主，完成游",
        "戏。反之，则游戏失败。",
    ]
    display_texts_page(screen_wel, texts, flip=False)

    font_wel = pygame.font.Font(FONT_PATH, 50)
    text_wel = font_wel.render("木木勇士，冲冲冲！", True, (235, 63, 47))
    screen_wel.blit(text_wel, [50, 300])

    pygame.display.flip()
    pygame.time.wait(4000)


def blood_loss(screen_loss):
    """
    if mumu get a wrong answer, mumu get blood loss. This prints the output tip
    @param screen_loss: the screen you are ready to make an output
    @return: none
    """
    font_loss = pygame.font.Font(FONT_PATH, 60)
    text_loss = font_loss.render("你个渣渣", True, RED)
    screen_loss.blit(text_loss, [100, 100])
    text_loss = font_loss.render("BLOOD -20~", True, RED)
    screen_loss.blit(text_loss, [100, 200])

    pygame.display.flip()
    pygame.time.wait(700)


def answer_right(screen_right):
    """
    if mumu get a right answer, mumu get a hint
    @param screen_right: the screen you are ready to make an output
    @return: none
    """
    font_right = pygame.font.Font(FONT_PATH, 60)
    text_right = font_right.render("You're Right~", True, RED)
    screen_right.blit(text_right, [100, 100])

    pygame.display.flip()
    pygame.time.wait(700)


def lose(screen_lose):
    """
    if you lose, you will see this
    @param screen_lose: the screen you are redy to make an output
    @return:
    """
    global done
    screen_lose.fill(WARM_GREY)

    pygame.mixer.music.fadeout(500)  # bgm music stop
    pygame.time.wait(500)

    pygame.mixer.music.load(LOSE_GAME_MUSIC_PATH)
    pygame.mixer.music.play()

    font_vic1 = pygame.font.Font(FONT_PATH, 60)
    font_vic2 = pygame.font.Font(FONT_PATH, 60)
    font_vic3 = pygame.font.Font(FONT_PATH, 60)
    font_vic4 = pygame.font.SysFont('TimesNewRoman', 60)
    text_vic1 = font_vic1.render("Oops~", True, BLACK)
    text_vic2 = font_vic2.render("看来你的火候还不够", True, BLACK)
    text_vic3 = font_vic3.render("那就陪公主一起挂科吧~", True, BLACK)
    text_vic4 = font_vic4.render("Press ESC to quit~ ", True, BLACK)

    screen_lose.blit(text_vic1, [100, 100])
    screen_lose.blit(text_vic2, [100, 200])
    screen_lose.blit(text_vic3, [100, 300])
    screen_lose.blit(text_vic4, [100, 400])

    pygame.display.flip()

    done = False
    while not done:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                done = True
            if event.type == pygame.KEYUP:
                if event.key == pygame.K_ESCAPE:
                    done = True
    pygame.quit()


def victory(screen_vic):
    """
    if you have won the game, you will see this hint
    @param screen_vic: the screen you are willing to make an output
    @return: none
    """
    global done
    screen_vic.fill(WARM_GREY)

    pygame.mixer.music.fadeout(1000)    # bgm music stop
    pygame.time.wait(500)

    pygame.mixer.music.load(VICTORY_MUSIC_PATH)
    pygame.mixer.music.play()

    font_vic1 = pygame.font.Font(FONT_PATH, 60)
    font_vic2 = pygame.font.Font(FONT_PATH, 60)
    font_vic3 = pygame.font.Font(FONT_PATH, 40)
    font_vic4 = pygame.font.Font(FONT_PATH, 40)
    font_vic5 = pygame.font.Font(FONT_PATH, 20)
    font_vic6 = pygame.font.Font(FONT_PATH, 20)
    text_vic1 = font_vic1.render("Congratulations!", True, BLACK)
    text_vic2 = font_vic2.render("你拯救了公主!", True, BLACK)
    text_vic3 = font_vic3.render("但由于你胆敢觊觎公主的美色，", True, BLACK)
    text_vic4 = font_vic4.render("国王决定将你处死...", True, BLACK)
    text_vic5 = font_vic5.render("（原来木木从来都只是国王的工具）", True, BLACK)
    text_vic6 = font_vic6.render("Press ESC to quit~ ", True, BLACK)

    screen_vic.blit(text_vic1, [100, 100])
    screen_vic.blit(text_vic2, [100, 200])
    screen_vic.blit(text_vic3, [100, 300])
    screen_vic.blit(text_vic4, [100, 350])
    screen_vic.blit(text_vic5, [100, 400])
    screen_vic.blit(text_vic6, [100, 470])

    pygame.display.flip()

    done = False
    while not done:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                done = True
            if event.type == pygame.KEYUP:
                if event.key == pygame.K_ESCAPE:
                    done = True
    pygame.quit()


# ------------------ Main Loop ------------------ #
def main():
    """
    the main loop funtion
    @return: none
    """
    # --------- PyGame Initialization --------- #
    pygame.init()

    size = (SCREEN_WIDTH, SCREEN_HEIGHT)   # window size
    screen = pygame.display.set_mode(size)
    pygame.display.set_caption("期末大作战")

    road_mat_12_16 = [[1]*16 for _ in range(12)]
    # entry point
    road_mat_12_16[0][1] = 0
    # final_point
    road_mat_12_16[-1][-2] = -1

    main_maze = Maze()
    main_maze.road_mat = road_mat_12_16
    player = Player(ROAD_THICKNESS, 0)

    road_list = pygame.sprite.Group()
    wall_list = pygame.sprite.Group()
    # generate the initial state
    main_maze.maze_generate(road_list, wall_list)

    wall_list_dfs = main_maze.dfs_maze_generate([0, 1], wall_list)
    while not wall_list_dfs:
        main_maze = Maze()
        road_list = pygame.sprite.Group()
        wall_list = pygame.sprite.Group()
        main_maze.road_mat = road_mat_12_16
        main_maze.maze_generate(road_list, wall_list)
        wall_list_dfs = main_maze.dfs_maze_generate([0, 1], wall_list)
        try:
            wall_list_dfs.draw(screen)
            pygame.display.flip()
        except Exception as e:
            print("wall_list_dfs error!" + str(Exception) + str(e))
            continue

    movingsprites = pygame.sprite.Group()
    movingsprites.add(player)

    # Teacher Section
    teacher_list = pygame.sprite.Group()
    for i in range(15):
        tea = Teacher()
        tea.rect.x = random.randrange(
            ROAD_THICKNESS, SCREEN_WIDTH - ROAD_THICKNESS, 50)
        tea.rect.y = random.randrange(
            ROAD_THICKNESS, SCREEN_HEIGHT - ROAD_THICKNESS, 50)
        teacher_list.add(tea)

    clock = pygame.time.Clock()

    # --------- BGM Loaded --------- #
    pygame.mixer.music.load(BACKGROUND_MUSIC_PATH)
    pygame.mixer.music.play()

    # --------- Figure Loaded --------- #
    xiaoxiao_image = pygame.image.load(PRINCESS_IMAGE_PATH).convert()

    # --------- Question Loaded --------- #
    wb = load_workbook(filename=QUESTION_WORKBOOK_PATH)
    ws = wb.get_sheet_by_name('Sheet1')

    # welcome
    welcome(screen)

    # --------- Main Loop --------- #
    global done
    global blood
    while not done:
        # --------- Event --------- #
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                done = True

            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_LEFT:
                    player.changespeed(-3, 0)
                if event.key == pygame.K_RIGHT:
                    player.changespeed(3, 0)
                if event.key == pygame.K_UP:
                    player.changespeed(0, -3)
                if event.key == pygame.K_DOWN:
                    player.changespeed(0, 3)

            if event.type == pygame.KEYUP:
                if event.key == pygame.K_LEFT:
                    player.changespeed(3, 0)
                if event.key == pygame.K_RIGHT:
                    player.changespeed(-3, 0)
                if event.key == pygame.K_UP:
                    player.changespeed(0, 3)
                if event.key == pygame.K_DOWN:
                    player.changespeed(0, -3)
                if event.key == pygame.K_ESCAPE:
                    done = True     # ESC key to quit the game

        # --------- Game Logic --------- #
        player.move(wall_list_dfs)

        # ask questions
        question_list = pygame.sprite.spritecollide(player, teacher_list, True)
        row_rand = random.randrange(1, 21, 1)
        for row in question_list:
            player.reset_speed()
            screen.fill(WARM_GREY)
            position = ('B' + str(row_rand))
            font_wb = pygame.font.Font(FONT_PATH, 15)
            text_wb = font_wb.render(ws[position].value, True, BLACK)

            position_a = ('C' + str(row_rand))
            font_option_a = pygame.font.Font(FONT_PATH, 15)
            text_option_a = font_option_a.render(
                ws[position_a].value, True, BLACK)
            position_b = ('D' + str(row_rand))
            font_option_b = pygame.font.Font(FONT_PATH, 15)
            text_option_b = font_option_b.render(
                ws[position_b].value, True, BLACK)
            position_c = ('E' + str(row_rand))
            font_option_c = pygame.font.Font(FONT_PATH, 15)
            text_option_c = font_option_c.render(
                ws[position_c].value, True, BLACK)
            position_d = ('F' + str(row_rand))
            font_option_d = pygame.font.Font(FONT_PATH, 15)
            text_option_d = font_option_d.render(
                ws[position_d].value, True, BLACK)

            screen.blit(text_wb, [50, 20])
            screen.blit(text_option_a, [50, 40])
            screen.blit(text_option_b, [50, 60])
            screen.blit(text_option_c, [50, 80])
            screen.blit(text_option_d, [50, 100])

            pygame.display.flip()
            # test if the answer is true
            ok = False
            right_answer = ws[('G'+str(row_rand))].value
            answer = ' '
            while not ok:
                for event in pygame.event.get():
                    if event.type == pygame.KEYUP:
                        if event.key == pygame.K_a:
                            answer = 'A'
                            ok = True
                        elif event.key == pygame.K_b:
                            answer = 'B'
                            ok = True
                        elif event.key == pygame.K_c:
                            answer = 'C'
                            ok = True
                        elif event.key == pygame.K_d:
                            answer = 'D'
                            ok = True
                        elif event.key == pygame.K_ESCAPE:
                            ok = True
                        else:
                            ok = False
            if answer == right_answer:
                answer_right(screen)
                # pygame.mixer.music.load(RIGHT_ANSWER_MUSIC_PATH)
                # pygame.mixer.music.play()
            else:
                blood -= 20
                blood_loss(screen)
                # pygame.mixer.music.load(WRONG_ANSWER_MUSIC_PATH)
                # pygame.mixer.music.play()

        if blood <= 0:
            lose(screen)

        if player.rect.top > SCREEN_HEIGHT - ROAD_THICKNESS \
                and player.rect.left >= SCREEN_WIDTH - 2 * ROAD_THICKNESS \
                and player.rect.right <= SCREEN_WIDTH and blood > 0:
            victory(screen)

        # --------- Game Graphics --------- #
        screen.fill(WARM_GREY)

        screen.blit(xiaoxiao_image,
                    (SCREEN_WIDTH - 2 * ROAD_THICKNESS,
                     SCREEN_HEIGHT-ROAD_THICKNESS))

        wall_list_dfs.draw(screen)
        teacher_list.draw(screen)
        movingsprites.draw(screen)

        # text on screen
        font = pygame.font.SysFont(FONT_PATH, 30)
        text = font.render("Blood: " + str(blood), True, RED)
        screen.blit(text, [0, 0])
        # --------- Refresh & Clock Set --------- #

        pygame.display.flip()   # fresh the screen

        clock.tick(60)  # the speed that the screen updates

    # --------- END of Game --------- #
    pygame.quit()


# ------------------ Debug Statement ------------------ #
if __name__ == "__main__":
    main()
